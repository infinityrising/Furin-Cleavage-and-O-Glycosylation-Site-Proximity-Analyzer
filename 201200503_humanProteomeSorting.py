##################################################
######################Dependencies################
##################################################
from openpyxl import Workbook
from operator import itemgetter
import numpy
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import os

#####################################################################################################
## Make a dictionary where the Fasta sequence name is the key, and the Fasta sequence is the value###
#####################################################################################################
inputAlignment = "modifieduniprot-keyword-Cleavage+on+pair+of+basic+residues-homo.fasta" # An alignment file produced by AliView.
inputAlignmentPath = "/Users/newumuser/Desktop/Coronavirus Python/Furin Cleavage and O-glycosylation/UniProt Homo Sapiens Swiss-Prot Reviewed Proteome 20350/5. Cleavage on Pair of Basic Residues/"
file = open(inputAlignmentPath + inputAlignment, "r")
lines = file.readlines() # The readline() function returns a list of lines in the file.
file.close()

keyCue=str('>s')
alignedSequences = {}
raw = []
i = 0
for line in lines:
	if keyCue in line:
		raw.append(line)
	else:
		modified = line.strip('\n')
		raw.append(modified)
longString = ''.join(raw)

oldNameKey = {} ###This is to ensure that we use the correct name in the final output. This is because the O-glycosylation prediction removes certain characters.
lessStrings = longString.split(keyCue)
del(lessStrings[0])

humanProteome = {} #Access the protein data as a dictionary of Protein :: Amino acid sequence
for each in lessStrings:
	modified = each.split('\n')
	humanProteome['>s'+modified[0]] = modified[1]

proteome = [] #a list of the proteome keys
for each in humanProteome.keys():
    proteome.append(each)

##############################################################################################
##########To generate a single fasta for PiTou prediction of furin cleavage sites#############
##############################################################################################
# proteomeFile = open('modified' + inputAlignment, 'w')
# k=0
# i=0
# for name in humanProteome:
# 	seq = humanProteome[name]
# 	proteomeFile.write(name + '\n')
# 	proteomeFile.write(seq + '\n')

##############################################################################################
#####Which sequences have at least one arginine predicted to be a site of furin cleavage?#####
##############################################################################################
inputAlignment = "modifieduniprot-keyword-Cleavage+on+pair+of+basic+residues-homo--FURINDATA.txt" # An alignment file produced by AliView.
inputAlignmentPath = "/Users/newumuser/Desktop/Coronavirus Python/Furin Cleavage and O-glycosylation/UniProt Homo Sapiens Swiss-Prot Reviewed Proteome 20350/5. Cleavage on Pair of Basic Residues/"
file = open(inputAlignmentPath + inputAlignment, "r")
lines = file.readlines() # The readline() function returns a list of lines in the file.
file.close()

PiTauPredictions = {} ####Create a dictionary with the PiTau results
lines = lines[1:]     ####Remove the header
for line in lines:
	
	stripped = line.split('\n')
	modified = stripped[0].split('\t')
	
	for correctName in proteome:

		if modified[0] in correctName:
			key = correctName

			if modified[11] == "Miaow!":
				pertinentData = [modified[1], modified[2], modified[3], modified[10], modified[11]]
				
                if key not in PiTauPredictions.keys():
					PiTauPredictions[key] = [pertinentData]
				
                else:
					PiTauPredictions[key].append(pertinentData)



#################################################################################################
###############################Parse the NetOGlyc4.0 data########################################
#################################################################################################
inputAlignment = "netoglyc4.0_cleavage dibasic residues.txt" # An alignment file produced by AliView.
inputAlignmentPath = "/Users/newumuser/Desktop/Coronavirus Python/Furin Cleavage and O-glycosylation/UniProt Homo Sapiens Swiss-Prot Reviewed Proteome 20350/5. Cleavage on Pair of Basic Residues/"
file = open(inputAlignmentPath + inputAlignment, "r")
lines = file.readlines() # The readline() function returns a list of lines in the file.
file.close()

fullOGlycDict = {} ####This paragraph of code takes the NetOGlyc4.0 data, changes the sequence name back to the original in the fasta,
lines = lines[5:] #####and creates a dictionary where key = seq name and value = [AA index, prediction], ...

for line in lines:
	
    stripped = line.strip('\n')
	positiveCuratedLine = stripped.split('\t')
	values = [positiveCuratedLine[3], positiveCuratedLine[5]]
	nameChange=positiveCuratedLine[0].split('_')
	
    for correctName in proteome:
		
        if nameChange[1] in correctName:
			positiveCuratedLine[0] = correctName
			
            if float(values[1]) < 0.5:
				pass
			
            else:
				
                if positiveCuratedLine[0] not in fullOGlycDict.keys():
					fullOGlycDict[correctName] = [values]
				
                else:
					fullOGlycDict[correctName].append(values)

#################################################################################################
######################Define a set of functions to perform our analyses##########################
#################################################################################################
factor = 15
def compareCleavageAndOGlycSite(cleavageSite, factor=15):
	cleavageSite = float(cleavageSite)
	#Todo: compare cleavage site and oglyc site and determine whether they are near each other
	def compare(oGlycSiteArr):
		oGlycSite = float(oGlycSiteArr[0])
		# print('Iterating', cleavageSite, oGlycSite, abs(cleavageSite-oGlycSite) <= factor)
		return abs(cleavageSite-oGlycSite) <= factor
	return compare

cleavageSite = PiTauPredictions[key][0][0] #77
OGlycSite = fullOGlycDict[key][0][0] #36

def uniqueOGlicSite():
    print('uniqueOGlicSite')
    hash = {}
    def f (v):
        was = not v[0] in hash
        print('v[0], was', v[0], was)
        hash[v[0]] = true
        return was
    return f


def cleanOGlycSite(PiTauPredictions, fullOGlycDict):
    newFullOGlycDict = {}
    for ptpKey in PiTauPredictions: #This finds a furin cleavage site
        ptpList = PiTauPredictions[ptpKey]
        for ptp in ptpList:
            cleavageSite = ptp[0] # 77
            i=0
            for ogsKey in fullOGlycDict: #This finds an oglycSite
                if not ogsKey in newFullOGlycDict:
                    newFullOGlycDict[ogsKey] = []

                if ogsKey != ptpKey:
                    continue

                if ogsKey == '>sp|P39905|GDNF_HUMAN Glial cell line-derived neurotrophic factor OS=Homo sapiens OX=9606 GN=GDNF PE=1 SV=1':
                    # print('key', ogsKey)
                    print('fullOGlycDict[ogsKey]', fullOGlycDict[ogsKey])
                    print('PiTauPredictions[ptpKey]', PiTauPredictions[ptpKey])
                # the filter function
                newFullOGlycDict[ogsKey].extend(filter(compareCleavageAndOGlycSite(cleavageSite), fullOGlycDict[ogsKey]))
                # newFullOGlycDict[ogsKey].append(cleavageSite)
                # print(newFullOGlycDict)
                i += 1
                # for listElement in newFullOGlycDict[ogsKey]:
                #     listElement.append(ptp[0])
                # newFullOGlycDict[ogsKey] = filter(uniqueOGlicSite(), newFullOGlycDict[ogsKey])

    # raise SystemExit
    return newFullOGlycDict

# print(fullOGlycDict)
# print('init', fullOGlycDict['>sp|P39905|GDNF_HUMAN Glial cell line-derived neurotrophic factor OS=Homo sapiens OX=9606 GN=GDNF PE=1 SV=1'])
print('result', cleanOGlycSite(PiTauPredictions, fullOGlycDict)['>sp|P39905|GDNF_HUMAN Glial cell line-derived neurotrophic factor OS=Homo sapiens OX=9606 GN=GDNF PE=1 SV=1'])
print('result', cleanOGlycSite(PiTauPredictions, fullOGlycDict))
#
# raise SystemExit

# import json

#################################################################################################
###############################Export data to an Excel File######################################
#################################################################################################

pirate = Workbook()
sheet = pirate.active
sheet.title = "Data Output"

cell = sheet.cell(row=1, column=1)
cell.value = "Sequence Name"
cell = sheet.cell(row=1, column=2)
cell.value = "Protein Length"
cell = sheet.cell(row=1, column=3)
cell.value = "# of PiTau Furin Cleavage Sites"
cell = sheet.cell(row=1, column=4)
cell.value = "# of predicted OGlyc Sites"
cell = sheet.cell(row=1, column=5)
cell.value = "# of predicted OGlyc Sites near a Furin Cleavage Site"
cell = sheet.cell(row=1, column=6)
cell.value = "Amino Acid #"
cell = sheet.cell(row=1, column=7)
cell.value = "Prediction Score"
cell = sheet.cell(row=1, column=8)
cell.value = "Amino Acid of Predicted Furin Cleavage Site"

print(cleanOGlycSite(PiTauPredictions, fullOGlycDict))

r=2
j=1
output = cleanOGlycSite(PiTauPredictions, fullOGlycDict)

for key in output:
    i=0
    if (j % 2) == 0:
        color = PatternFill(start_color="CCCC99", fill_type = "solid")
    else:
        color = PatternFill(start_color="FFCC99", fill_type = "solid")
    if key not in PiTauPredictions.keys():
        j+=1
        pass
    else:
        seqDataSet = output[key]
        name = key
        length = len(humanProteome[key])
        fcsNumber = len(PiTauPredictions[key])
        ogsNumber = len(fullOGlycDict[key])

        #Todo to acquire each index of each furin cleavage site
        fcsIndicesList = []
        for each in PiTauPredictions[key]:
            fcsChosenInfo = [each[0], each[3]]
            fcsIndicesList.append(fcsChosenInfo)
        print(fcsIndicesList)

        #Todo acquire # of predicted OGlyc Sites near a furin cleavage site
        uniqueIndexList = []
        for each in output[key]:
            if each not in uniqueIndexList:
                uniqueIndexList.append(each)
        ogsNearFurin = len(uniqueIndexList)
        print(ogsNearFurin, uniqueIndexList)

        cell = sheet.cell(row=r, column=1)
        cell.value = name
        cell.fill = color
        cell = sheet.cell(row=r, column=2)
        cell.value = length
        cell.fill = color
        cell = sheet.cell(row=r, column=3)
        cell.value = fcsNumber
        cell.fill = color
        cell = sheet.cell(row=r, column=4)
        cell.value = ogsNumber
        cell.fill = color
        cell = sheet.cell(row=r, column=5)
        cell.value = ogsNearFurin
        cell.fill = color
        if ogsNearFurin == 0:
            cell = sheet.cell(row=r, column=6)
            cell.value = 'N/A'
            cell.fill = color
            cell = sheet.cell(row=r, column=7)
            cell.value = 'N/A'
            cell.fill = color
            r+=1
        else:
            cell = sheet.cell(row=r, column=6)
            cell.value = uniqueIndexList[0][0]
            cell.fill = color
            cell = sheet.cell(row=r, column=7)
            cell.value = uniqueIndexList[0][1]
            cell.fill = color

            for each in fcsIndicesList:
                pertinentDataList = []
                pertinentData = each[0] + ' (' + each[1] + ')'
                if abs(float(each[0])-float(uniqueIndexList[0][0])) <= factor:
                    pertinentData = each[0] + ' (' + each[1] + ')'
                    pertinentDataList.append(pertinentData)
                for each in pertinentDataList:
                    cell = sheet.cell(row=r, column=8)
                    # cell.value = each + '\015'
                    # cell.style.alignment.wrap_text = True
                    cell.value = each + '\n'
                    cell.fill = color

            i = 1
            for score in uniqueIndexList[1:]:
                glycAminoAcidIndex = score[0]
                glycPrediction = score[1]
                r += 1
                cell = sheet.cell(row=r, column=1)
                cell.value = name
                cell.fill = color
                cell = sheet.cell(row=r, column=2)
                cell.value = length
                cell.fill = color
                cell = sheet.cell(row=r, column=3)
                cell.value = fcsNumber
                cell.fill = color
                cell = sheet.cell(row=r, column=4)
                cell.value = ogsNumber
                cell.fill = color
                cell = sheet.cell(row=r, column=5)
                cell.value = ogsNearFurin
                cell.fill = color
                cell = sheet.cell(row=r, column=6)
                cell.value = glycAminoAcidIndex
                cell.fill = color
                cell = sheet.cell(row=r, column=7)
                cell.value = glycPrediction
                cell.fill = color

                #Iterating the furin cleavage sites
                pertinentDataList = []
                for each in fcsIndicesList:
                    pertinentData = each[0] + ' (' + each[1] + ')'
                    if abs(float(each[0])-float(glycAminoAcidIndex)) <= factor:
                        pertinentData = each[0] + ' (' + each[1] + ')'
                        pertinentDataList.append(pertinentData)
                    cell = sheet.cell(row=r, column=8)
                    if len(pertinentDataList) == 0:
                        cell.value = "N/A"
                    else:
                        z=0
                        for each in pertinentDataList:
                            cell.value = pertinentDataList[z]
                            cell.alignment = Alignment(wrapText=True)
                            z+=1
                    cell.fill = color

                i += 1
            r += 1
    j+=1

newInputPath = '/Users/newumuser/Desktop/Coronavirus Python/Furin Cleavage and O-glycosylation/UniProt Homo Sapiens Swiss-Prot Reviewed Proteome 20350/'
pirate.save(newInputPath + "analysis" + ".xlsx")

#
# raise SystemExit
#
# #####Output to a text file
# def resultPrint(newFullOGlycDict):
#     resultFile = open('results.txt', 'a')
#     for key in cleanOGlycSite(PiTauPredictions, fullOGlycDict):
#         value = cleanOGlycSite(PiTauPredictions, fullOGlycDict)[key]
#         if not value:
#             continue
#         resultFile.write(key + '\n')
#         resultFile.write('Protein length: ' + str(len(humanProteome[key])) + ' amino acids' + '\n')
#         resultFile.write('Number of predicted furin cleavage sites: ' + str(len(PiTauPredictions[key])) + '\n')
#         resultFile.write('Number of S/T predicted to be O-glycosylated (>0.5): ' + str(len(fullOGlycDict[key])) + '\n')
#         resultFile.write('Number of O-glycosylation sites near a furin cleavage site: ' + str(len(value)) + '\n')
#         resultFile.write('These O-glycosylation sites (prediction score) are close to a Furin cleavage site:' + '\n')
#         for ogsNearFurin in value:
#             resultFile.write('\t' + 'Amino acid: ' + ogsNearFurin[0] + ' (' + str(ogsNearFurin[1]) + ')' + '\n')
#         resultFile.write('**If an amino acid appears twice then it falls within the proximity of two furin cleavage sites' + '\n' + '\n' + '\n')
#     resultFile.close()
#
# print(resultPrint(cleanOGlycSite(PiTauPredictions, fullOGlycDict)))
#
# print(len(cleanOGlycSite(PiTauPredictions, fullOGlycDict)))
# raise SystemExit
#
# # proteomeFile = open('modified' + inputAlignment, 'w')
# # k=0
# # i=0
# # for name in humanProteome:
# # 	seq = humanProteome[name]
# # 	proteomeFile.write(name + '\n')
# # 	proteomeFile.write(seq + '\n')
# #
# raise SystemExit
